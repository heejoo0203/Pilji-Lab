from __future__ import annotations

from pathlib import Path


def write_result_workbook(
    *,
    output_path: Path,
    headers: list[str],
    original_rows: list[list[str]],
    year_columns: list[str],
    row_results: list[dict[str, object]],
) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "조회결과"

    extra_headers = [f"공시지가_{year}" for year in year_columns] + ["조회상태", "오류사유", "주소요약", "PNU"]
    all_headers = [*headers, *extra_headers]
    sheet.append(all_headers)

    for col in range(1, len(all_headers) + 1):
        cell = sheet.cell(row=1, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="EEF4FF", end_color="EEF4FF", fill_type="solid")
        sheet.column_dimensions[cell.column_letter].width = 16

    for idx, raw_row in enumerate(original_rows):
        result = row_results[idx] if idx < len(row_results) else {}
        year_values: dict[str, str] = result.get("year_values", {}) if isinstance(result.get("year_values"), dict) else {}
        status = str(result.get("status", "오류"))
        error = str(result.get("error_message", ""))
        summary = str(result.get("address_summary", ""))
        pnu = str(result.get("pnu", ""))

        appended_year_values = [year_values.get(year, "") for year in year_columns]
        sheet.append([*raw_row, *appended_year_values, status, error, summary, pnu])

    workbook.save(output_path)
