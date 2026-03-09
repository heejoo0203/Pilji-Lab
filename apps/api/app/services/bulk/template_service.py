from __future__ import annotations

from io import BytesIO

from app.services.bulk.constants import RECOMMENDED_JIBUN, RECOMMENDED_ROAD


def build_template_workbook_bytes() -> bytes:
    from openpyxl import Workbook

    workbook = Workbook()
    guide = workbook.active
    guide.title = "가이드"
    _write_guide_sheet(guide)

    jibun_sheet = workbook.create_sheet("지번_예시")
    _write_example_sheet(jibun_sheet, RECOMMENDED_JIBUN, ["서울특별시", "강남구", "도곡동", "일반", "970", "0"])

    road_sheet = workbook.create_sheet("도로명_예시")
    _write_example_sheet(road_sheet, RECOMMENDED_ROAD, ["서울특별시", "강남구", "도곡로", "21", "0"])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _write_guide_sheet(sheet: object) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.worksheet import Worksheet

    ws = sheet if isinstance(sheet, Worksheet) else None
    if ws is None:
        return

    ws["A1"] = "필지랩 파일조회 표준 양식 안내"
    ws["A1"].font = Font(size=14, bold=True)
    ws.merge_cells("A1:D1")

    ws["A3"] = "필수 공통 컬럼"
    ws["A3"].font = Font(bold=True)
    ws["A4"] = "주소유형 (지번/도로명)"

    ws["A6"] = "지번 권장 컬럼"
    ws["A6"].font = Font(bold=True)
    ws["A7"] = ", ".join(RECOMMENDED_JIBUN)

    ws["A9"] = "도로명 권장 컬럼"
    ws["A9"].font = Font(bold=True)
    ws["A10"] = ", ".join(RECOMMENDED_ROAD)

    ws["A12"] = "주의사항"
    ws["A12"].font = Font(bold=True)
    ws["A13"] = "열 순서가 달라도 자동 매핑되지만, 권장 컬럼명을 지키면 정확도가 높아집니다."
    ws["A14"] = "최대 10,000행까지 처리됩니다."
    ws["A15"] = "결과 파일에는 연도별 공시지가 컬럼이 최신순으로 추가됩니다."

    ws.column_dimensions["A"].width = 88
    for row in (3, 6, 9, 12):
        ws[f"A{row}"].fill = PatternFill(start_color="E8F1FF", end_color="E8F1FF", fill_type="solid")
        ws[f"A{row}"].alignment = Alignment(horizontal="left")


def _write_example_sheet(sheet: object, headers: list[str], values: list[str]) -> None:
    from openpyxl.styles import Font, PatternFill
    from openpyxl.worksheet.worksheet import Worksheet

    ws = sheet if isinstance(sheet, Worksheet) else None
    if ws is None:
        return

    ws.append(["주소유형", *headers])
    ws.append([_mode_for_headers(headers), *values])

    for col_idx in range(1, len(headers) + 3):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="F2F7FF", end_color="F2F7FF", fill_type="solid")
        ws.column_dimensions[cell.column_letter].width = 17


def _mode_for_headers(headers: list[str]) -> str:
    return "지번" if "읍면동" in headers else "도로명"
