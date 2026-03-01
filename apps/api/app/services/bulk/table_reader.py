from __future__ import annotations

import csv
import io
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path


@dataclass(slots=True)
class TabularData:
    headers: list[str]
    rows: list[list[str]]


def load_tabular_data(file_path: Path) -> TabularData:
    ext = file_path.suffix.lower()
    if ext == ".xlsx":
        return _load_xlsx(file_path)
    if ext == ".xls":
        return _load_xls(file_path)
    if ext == ".csv":
        return _load_csv(file_path)
    raise ValueError("지원하지 않는 파일 형식입니다. (.xlsx, .xls, .csv)")


def _load_xlsx(file_path: Path) -> TabularData:
    from openpyxl import load_workbook

    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    sheet = workbook.active
    try:
        rows = sheet.iter_rows(values_only=True)
        header_raw = _take_first_non_empty_row(rows)
        if header_raw is None:
            return TabularData(headers=[], rows=[])

        headers = _normalize_headers([_cell_to_text(v) for v in header_raw])
        body_rows: list[list[str]] = []
        for row in rows:
            cells = [_cell_to_text(v) for v in row]
            if _is_empty_row(cells):
                continue
            body_rows.append(_fit_row(cells, len(headers)))
        return TabularData(headers=headers, rows=body_rows)
    finally:
        workbook.close()


def _load_xls(file_path: Path) -> TabularData:
    import xlrd

    workbook = xlrd.open_workbook(file_path.as_posix())
    sheet = workbook.sheet_by_index(0)
    header_raw: list[str] | None = None
    start_index = 0
    for row_idx in range(sheet.nrows):
        row_values = [_cell_to_text(sheet.cell_value(row_idx, col_idx)) for col_idx in range(sheet.ncols)]
        if _is_empty_row(row_values):
            continue
        header_raw = row_values
        start_index = row_idx + 1
        break

    if header_raw is None:
        return TabularData(headers=[], rows=[])

    headers = _normalize_headers(header_raw)
    body_rows: list[list[str]] = []
    for row_idx in range(start_index, sheet.nrows):
        row_values = [_cell_to_text(sheet.cell_value(row_idx, col_idx)) for col_idx in range(sheet.ncols)]
        if _is_empty_row(row_values):
            continue
        body_rows.append(_fit_row(row_values, len(headers)))
    return TabularData(headers=headers, rows=body_rows)


def _load_csv(file_path: Path) -> TabularData:
    raw = file_path.read_bytes()
    text = _decode_csv_bytes(raw)
    reader = csv.reader(io.StringIO(text))
    header_raw: list[str] | None = None
    body_rows: list[list[str]] = []
    for row in reader:
        cells = [_cell_to_text(v) for v in row]
        if header_raw is None:
            if _is_empty_row(cells):
                continue
            header_raw = cells
            continue
        if _is_empty_row(cells):
            continue
        body_rows.append(cells)

    if header_raw is None:
        return TabularData(headers=[], rows=[])

    headers = _normalize_headers(header_raw)
    fitted = [_fit_row(row, len(headers)) for row in body_rows]
    return TabularData(headers=headers, rows=fitted)


def _take_first_non_empty_row(rows: object) -> tuple[object, ...] | None:
    for row in rows:
        if row is None:
            continue
        values = [_cell_to_text(v) for v in row]
        if not _is_empty_row(values):
            return row
    return None


def _normalize_headers(headers: list[str]) -> list[str]:
    normalized: list[str] = []
    duplicates: dict[str, int] = {}
    for idx, header in enumerate(headers):
        base = header.strip() or f"column_{idx + 1}"
        count = duplicates.get(base, 0) + 1
        duplicates[base] = count
        if count > 1:
            normalized.append(f"{base}_{count}")
        else:
            normalized.append(base)
    return normalized


def _fit_row(row: list[str], width: int) -> list[str]:
    if len(row) > width:
        return row[:width]
    if len(row) < width:
        return row + [""] * (width - len(row))
    return row


def _is_empty_row(cells: list[str]) -> bool:
    return all(not str(cell).strip() for cell in cells)


def _cell_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).rstrip("0").rstrip(".")
    if isinstance(value, int):
        return str(value)
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value).strip()


def _decode_csv_bytes(raw: bytes) -> str:
    for encoding in ("utf-8-sig", "cp949", "euc-kr"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="ignore")
